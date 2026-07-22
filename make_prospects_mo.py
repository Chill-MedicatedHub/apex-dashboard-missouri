"""
make_prospects_mo.py — build prospects.json for the Missouri Prospecting tab
(New / potential buyers), sourced from the state regulator.

Missouri's Division of Cannabis Regulation publishes the licensed-dispensary
list as a downloadable Excel file (public URL — no login, no token). This
script downloads it (plus the micro-dispensary list), extracts every licensed
dispensary with its DIS###### license number, and writes prospects.json.

The dashboard then compares that list against your sales book by license number
and shows only the dispensaries you have not yet invoiced.

Runs anywhere with internet — locally or in GitHub Actions. No credentials.

Usage:
    python make_prospects_mo.py            # writes prospects.json here
    python make_prospects_mo.py out.json   # custom output path

Requires: requests, openpyxl  (pip install requests openpyxl)
"""

import io, json, re, sys
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

# Missouri DCR licensed-facility spreadsheets (public).
URLS = [
    "https://health.mo.gov/safety/cannabis/xls/licensed-dispensary-facilities-508.xlsx",
    "https://health.mo.gov/safety/cannabis/xls/micro-licensed-dispensary-facilities.xlsx",
]
OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).parent / "prospects.json"

# Header keywords -> which field we map the column to. First match wins per column.
HEADER_MAP = [
    ("license", "license"),      # "License Number", "License #", "License ID"
    ("dba", "dba"),              # "DBA", "DBA/Trade Name"
    ("trade", "dba"),
    ("entity", "name"),          # "Entity Name", "Legal Entity Name"
    ("legal", "name"),
    ("business", "name"),
    ("facility name", "name"),
    ("name", "name"),            # generic fallback (checked last among names)
    ("address", "address"),
    ("street", "address"),
    ("city", "city"),
    ("zip", "zip"),
    ("postal", "zip"),
]

LIC_RE = re.compile(r"\bDIS[-\s]?\d{3,}\b", re.I)   # Missouri dispensary license


def _norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


def _classify(header):
    """Return the field name a header cell maps to, or None."""
    h = _norm(header).lower()
    if not h:
        return None
    for kw, field in HEADER_MAP:
        if kw in h:
            return field
    return None


def _find_header_row(ws, scan=12):
    """Locate the row that contains the column headers by finding the first row
    whose cells map to at least a license column + one name/city column."""
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=scan, values_only=True), start=1):
        fields = {}
        for c, val in enumerate(row):
            f = _classify(val)
            if f and f not in fields:
                fields[f] = c
        if "license" in fields and (("name" in fields) or ("dba" in fields) or ("city" in fields)):
            return r, fields
    return None, None


def _clean_license(v):
    m = LIC_RE.search(str(v or ""))
    if not m:
        return ""
    return re.sub(r"[-\s]", "", m.group(0)).upper()


def parse_sheet(content, src):
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    records = []
    for ws in wb.worksheets:
        hdr_row, cols = _find_header_row(ws)
        if not cols:
            continue
        li, ni, di, ci, ai, zi = (cols.get("license"), cols.get("name"),
                                   cols.get("dba"), cols.get("city"),
                                   cols.get("address"), cols.get("zip"))
        for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            def cell(idx):
                return _norm(row[idx]) if (idx is not None and idx < len(row)) else ""
            lic = _clean_license(cell(li))
            if not lic:
                continue
            # Prefer the DBA / trade name (what the store is actually called);
            # fall back to the legal entity name.
            name = cell(di) or cell(ni) or "—"
            records.append({
                "name": name, "license": lic, "type": "Dispensary", "status": "Active",
                "city": cell(ci), "address": cell(ai), "zip": cell(zi),
            })
    print(f"  {src.split('/')[-1]}: {len(records)} dispensary rows")
    return records


def main():
    all_recs, seen = [], set()
    for url in URLS:
        try:
            r = requests.get(url, timeout=60,
                             headers={"User-Agent": "chill-prospects"})
        except requests.RequestException as e:
            print(f"  WARN: could not download {url} ({e}); skipping.")
            continue
        if r.status_code != 200:
            print(f"  WARN: {url} returned {r.status_code}; skipping.")
            continue
        try:
            recs = parse_sheet(r.content, url)
        except Exception as e:
            print(f"  WARN: could not parse {url} ({e}); skipping.")
            continue
        for rec in recs:
            if rec["license"] in seen:
                continue
            seen.add(rec["license"])
            all_recs.append(rec)

    if not all_recs:
        sys.exit("ERROR: no dispensary records parsed. The Missouri spreadsheet layout "
                 "may have changed — check the header names and update HEADER_MAP.")

    all_recs.sort(key=lambda x: (x["city"], x["name"]))
    payload = {
        "as_of": datetime.now(timezone.utc).date().isoformat(),
        "source": "Missouri DCR licensed-dispensary facilities (health.mo.gov)",
        "state": "MO",
        "count": len(all_recs),
        "records": all_recs,
    }
    OUT.write_text(json.dumps(payload, indent=2))
    print(f"Wrote {OUT}: {len(all_recs)} licensed dispensaries.")
    print("The dashboard drops the ones you've already invoiced and shows the rest.")


if __name__ == "__main__":
    main()
